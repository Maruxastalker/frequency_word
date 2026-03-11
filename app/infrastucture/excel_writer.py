"""
Инфраструктурный сервис записи в Excel.
"""

import tempfile

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from app.domain.models.word_frequency import FrequencyReport


class ExcelReportWriter:
    """Записывает FrequencyReport в xlsx-файл."""

    HEADER_FONT = Font(bold=True, size=12, color="FFFFFF")
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def write(self, report: FrequencyReport) -> str:
        """
        Записать отчёт во временный xlsx-файл.
        Возвращает путь к файлу.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Частотный анализ"

        # Заголовки
        headers = [
            "Словоформа (лемма)",
            "Кол-во во всём документе",
            "Кол-во в каждой строке"
        ]
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.THIN_BORDER

        # Данные
        entries = report.get_sorted_entries()
        
        for row_idx, wf in enumerate(entries, start=2):
            cell_lemma = ws.cell(row=row_idx, column=1, value=wf.lemma)
            cell_lemma.border = self.THIN_BORDER

            cell_total = ws.cell(row=row_idx, column=2, value=wf.total_count)
            cell_total.border = self.THIN_BORDER
            cell_total.alignment = Alignment(horizontal="center")

            line_dist = wf.get_line_distribution_str(report.total_lines)
            cell_lines = ws.cell(row=row_idx, column=3, value=line_dist)
            cell_lines.border = self.THIN_BORDER

        # Ширина столбцов
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 60

        # Сохраняем
        tmp = tempfile.NamedTemporaryFile(
            suffix=".xlsx",
            prefix="frequency_report_",
            delete=False
        )
        tmp.close()
        
        wb.save(tmp.name)
        wb.close()

        return tmp.name